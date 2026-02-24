import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import calendar

#DEF(1) QUEBRA DE LINHA MANUAL/ESTILO WORD PARA EXCEL CUIDA
# =========================
def quebrar_texto_word(texto: str, max_chars: int = 52) -> str:
    if not texto:
        return ""

    texto = str(texto).replace("\r\n", "\n").replace("\r", "\n").strip()

    if "\n" in texto:
        return texto

    palavras = texto.split()
    linhas = []
    linha = ""

    for p in palavras:
        if not linha:
            linha = p
        elif len(linha) + 1 + len(p) <= max_chars:
            linha += " " + p
        else:
            linhas.append(linha)
            linha = p

    if linha:
        linhas.append(linha)

    return "\n".join(linhas)

# DEF(1) LEITURA DAS ENTRADAS DO PDF DOC. CUIDA
# =====================================================================================================================
def ler_entradas_pdf(caminho_pdf: str) -> dict:
    texto_completo = ""
    dados_entradas = {}
    data_atual = None

    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                texto_completo += texto + "\n"

    for linha in texto_completo.split("\n"):
        if "Data de Rateio" in linha:
            partes = linha.split(":")
            data_atual = partes[1].strip()
            if data_atual not in dados_entradas:
                dados_entradas[data_atual] = []
        else:
            partes = linha.split()
            if len(partes) > 8 and partes[0].isdigit():
                guia = "SICASE - " + partes[6]
                emolumento = partes[8]
                if data_atual:
                    dados_entradas[data_atual].append({
                        "descricao": guia,
                        "entrada": emolumento,
                        "saida": 0.0
                    })

    return dados_entradas

# DEF(2) LEITURA DAS SAIDAS EXCEL DOC.
# ===========================================================================================================================
def ler_saidas_excel(caminho_excel_saidas: str) -> dict:
    wb_saida = load_workbook(caminho_excel_saidas, data_only=True)
    ws_saida = wb_saida.active

    dados_saidas = {}

    for row in ws_saida.iter_rows(min_row=2, values_only=True):
        data = row[0]
        descricao = row[5]
        valor = row[17]  

        if isinstance(data, datetime) and isinstance(valor, (int, float)) and descricao:
            data_str = data.strftime("%d/%m/%Y")
            descricao_str = str(descricao).replace("\r\n", "\n").replace("\r", "\n").strip()

            if data_str not in dados_saidas:
                dados_saidas[data_str] = []

            dados_saidas[data_str].append({
                "descricao": descricao_str,
                "entrada": 0.0,
                "saida": float(valor)
            })

    return dados_saidas



# DEF(3) RELACIONAR E JUNTAR POR DATA
# ==============================================================================================================================
def juntar_movimentos_por_data(entradas: dict, saidas: dict) -> dict:
    todas_datas = set(entradas.keys()) | set(saidas.keys())
    dados = {}
    for data in todas_datas:
        movs = []
        movs.extend(entradas.get(data, []))
        movs.extend(saidas.get(data, []))
        dados[data] = movs
    return dados

# DEF(4) GERAR O LIVRO DIÁRIO AND RESUMO NO FINAL
# ===========================================================================================================
def gerar_livro_diario(dados: dict, nome_saida: str, max_chars_saida: int = 52):
    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMO"

    estilo_colorido = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fonte_padrao = Font(name='Arial', size=12, bold=False)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    borda_fina = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    borda_esquerda_vazia = Border(left=Side(style='thin'))
    borda_direita_vazia = Border(right=Side(style='thin'))
    sem_borda = Border()

    alinhamento_desc_entrada = Alignment(horizontal="left", vertical="center", wrap_text=False)
    alinhamento_desc_saida = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def aplicar_estilo_contabil(celula):
        celula.number_format = '_-* #,##0.00_-; \\-* #,##0.00_-;_-* "-"??_-;_-@_-'
        celula.border = borda_fina
        celula.font = fonte_padrao
        celula.alignment = alinhamento_centro

    primeira_data = list(dados.keys())[0]
    _, mes, ano = map(int, primeira_data.split("/"))
    num_dias = calendar.monthrange(ano, mes)[1]

    saldo_atual = 0.0
    linha_excel_atual = 1
    linha_saldo_dia_anterior = 1

    resumo_mensal = []

    cabecalhos = ["DATA", "DESCRIÇÃO DO LANÇAMENTO", "ENTRADA", "SAÍDA", "SALDO"]

    def escrever_cabecalho():
        nonlocal linha_excel_atual
        for col, texto in enumerate(cabecalhos, 1):
            celula = ws.cell(row=linha_excel_atual, column=col, value=texto)
            celula.font = fonte_padrao
            celula.alignment = alinhamento_centro
            celula.border = borda_fina
        ws.row_dimensions[linha_excel_atual].height = 15
        linha_excel_atual += 1

    for dia in range(1, num_dias + 1):
        data = f"{dia:02d}/{mes:02d}/{ano}"
        movimentos = dados.get(data, [])

        escrever_cabecalho()

        entrada_dia = 0.0
        saida_dia = 0.0
        primeira_linha_do_dia = True

        num_linhas_vazias = 0 if len(movimentos) > 0 else 10

        for m in movimentos:
            if isinstance(m["entrada"], str):
                entrada = float(m["entrada"].replace(".", "").replace(",", "."))
            else:
                entrada = float(m["entrada"])

            saida = float(m["saida"])

            entrada_dia += entrada
            saida_dia += saida
            saldo_atual += (entrada - saida)

            c_data = ws.cell(row=linha_excel_atual, column=1, value=data)
            c_data.border = borda_fina
            c_data.font = fonte_padrao
            c_data.alignment = alinhamento_centro

            desc_original = m["descricao"]
            if saida > 0:
                desc = quebrar_texto_word(desc_original, max_chars=max_chars_saida)
                c_desc = ws.cell(row=linha_excel_atual, column=2, value=desc)
                c_desc.alignment = alinhamento_desc_saida
                qtd_linhas = desc.count("\n") + 1
                ws.row_dimensions[linha_excel_atual].height = max(15, 15 * qtd_linhas)
            else:
                c_desc = ws.cell(row=linha_excel_atual, column=2, value=desc_original)
                c_desc.alignment = alinhamento_desc_entrada
                ws.row_dimensions[linha_excel_atual].height = 15

            c_desc.border = borda_fina
            c_desc.font = fonte_padrao

            c_entrada = ws.cell(row=linha_excel_atual, column=3, value=entrada)
            c_saida = ws.cell(row=linha_excel_atual, column=4, value=saida)

            c_saldo = ws.cell(row=linha_excel_atual, column=5)
            if primeira_linha_do_dia:
                c_saldo.value = f"=E{linha_saldo_dia_anterior}+C{linha_excel_atual}-D{linha_excel_atual}"
                primeira_linha_do_dia = False
            else:
                c_saldo.value = f"=E{linha_excel_atual-1}+C{linha_excel_atual}-D{linha_excel_atual}"

            for c in [c_entrada, c_saida, c_saldo]:
                aplicar_estilo_contabil(c)

            linha_excel_atual += 1

        for _ in range(num_linhas_vazias):
            c_data_v = ws.cell(row=linha_excel_atual, column=1, value=data)
            c_data_v.border = borda_fina
            c_data_v.font = fonte_padrao
            c_data_v.alignment = alinhamento_centro

            c_desc_v = ws.cell(row=linha_excel_atual, column=2, value=None)
            c_desc_v.border = borda_fina
            c_desc_v.font = fonte_padrao
            c_desc_v.alignment = alinhamento_desc_entrada

            c_entrada = ws.cell(row=linha_excel_atual, column=3, value=0)
            c_saida = ws.cell(row=linha_excel_atual, column=4, value=0)

            c_saldo = ws.cell(row=linha_excel_atual, column=5)
            if primeira_linha_do_dia:
                c_saldo.value = f"=E{linha_saldo_dia_anterior}+C{linha_excel_atual}-D{linha_excel_atual}"
                primeira_linha_do_dia = False
            else:
                c_saldo.value = f"=E{linha_excel_atual-1}+C{linha_excel_atual}-D{linha_excel_atual}"

            for c in [c_entrada, c_saida, c_saldo]:
                aplicar_estilo_contabil(c)

            ws.row_dimensions[linha_excel_atual].height = 15
            linha_excel_atual += 1

        ws.merge_cells(start_row=linha_excel_atual, start_column=1, end_row=linha_excel_atual, end_column=2)
        c_label = ws.cell(row=linha_excel_atual, column=1, value="Saldo do Dia")
        c_label.alignment = alinhamento_centro
        c_label.font = fonte_padrao

        for col in range(1, 6):
            celula = ws.cell(row=linha_excel_atual, column=col)
            celula.border = borda_fina
            celula.font = fonte_padrao
            celula.alignment = alinhamento_centro
            if col >= 3:
                if col == 3:
                    celula.value = entrada_dia
                if col == 4:
                    celula.value = saida_dia
                if col == 5:
                    celula.value = saldo_atual
                    linha_saldo_dia_anterior = linha_excel_atual
                aplicar_estilo_contabil(celula)

        ws.row_dimensions[linha_excel_atual].height = 15
        linha_excel_atual += 1

        for col in range(1, 6):
            celula_vazia = ws.cell(row=linha_excel_atual, column=col)
            celula_vazia.fill = estilo_colorido
            if col == 1:
                celula_vazia.border = borda_esquerda_vazia
            elif col == 5:
                celula_vazia.border = borda_direita_vazia
            else:
                celula_vazia.border = sem_borda

        ws.row_dimensions[linha_excel_atual].height = 15
        linha_excel_atual += 1

        status = "Movimento do dia" if (entrada_dia != 0 or saida_dia != 0) else "Sem Movimento"
        resumo_mensal.append((data, status, entrada_dia, saida_dia))

    # ================================================================================================================
    # RESUMO NO FINAL DOS DIAS/ESTILO
    ws.merge_cells(start_row=linha_excel_atual, start_column=1, end_row=linha_excel_atual, end_column=5)
    c_titulo = ws.cell(row=linha_excel_atual, column=1, value="RESUMO")
    c_titulo.font = fonte_padrao
    c_titulo.alignment = alinhamento_centro

    for col in range(1, 6):
        ws.cell(row=linha_excel_atual, column=col).border = borda_fina
        ws.cell(row=linha_excel_atual, column=col).font = fonte_padrao
        ws.cell(row=linha_excel_atual, column=col).alignment = alinhamento_centro

    ws.row_dimensions[linha_excel_atual].height = 15
    linha_excel_atual += 1

    saldo_resumo = 0.0

    for (data, status, ent, sai) in resumo_mensal:
        cA = ws.cell(row=linha_excel_atual, column=1, value=data)
        cA.border = borda_fina
        cA.font = fonte_padrao
        cA.alignment = alinhamento_centro

        cB = ws.cell(row=linha_excel_atual, column=2, value=status)
        cB.border = borda_fina
        cB.font = fonte_padrao
        cB.alignment = Alignment(horizontal="left", vertical="center")

        cC = ws.cell(row=linha_excel_atual, column=3, value=ent)
        aplicar_estilo_contabil(cC)

        cD = ws.cell(row=linha_excel_atual, column=4, value=sai)
        aplicar_estilo_contabil(cD)

        saldo_resumo += (ent - sai)
        cE = ws.cell(row=linha_excel_atual, column=5, value=saldo_resumo)
        aplicar_estilo_contabil(cE)

        for col in range(1, 6):
            ws.cell(row=linha_excel_atual, column=col).border = borda_fina
            ws.cell(row=linha_excel_atual, column=col).font = fonte_padrao

        ws.row_dimensions[linha_excel_atual].height = 15
        linha_excel_atual += 1

    ws.merge_cells(start_row=linha_excel_atual, start_column=1, end_row=linha_excel_atual, end_column=2)
    c_total_lbl = ws.cell(row=linha_excel_atual, column=1, value="TOTAL")
    c_total_lbl.font = fonte_padrao
    c_total_lbl.alignment = alinhamento_centro
    c_total_lbl.border = borda_fina

    total_ent = sum(x[2] for x in resumo_mensal)
    total_sai = sum(x[3] for x in resumo_mensal)

    cC = ws.cell(row=linha_excel_atual, column=3, value=total_ent)
    cD = ws.cell(row=linha_excel_atual, column=4, value=total_sai)
    cE = ws.cell(row=linha_excel_atual, column=5, value=saldo_resumo)

    aplicar_estilo_contabil(cC)
    aplicar_estilo_contabil(cD)
    aplicar_estilo_contabil(cE)

    for col in range(1, 6):
        ws.cell(row=linha_excel_atual, column=col).border = borda_fina

    ws.row_dimensions[linha_excel_atual].height = 15
    linha_excel_atual += 1

    for col in range(1, 6):
        cel = ws.cell(row=linha_excel_atual, column=col)
        cel.fill = estilo_colorido
        if col == 1:
            cel.border = borda_esquerda_vazia
        elif col == 5:
            cel.border = borda_direita_vazia
        else:
            cel.border = sem_borda

    ws.row_dimensions[linha_excel_atual].height = 15
    linha_excel_atual += 1

    ws.column_dimensions['A'].width = 14.72
    ws.column_dimensions['B'].width = 65.72
    ws.column_dimensions['C'].width = 14.72
    ws.column_dimensions['D'].width = 14.72
    ws.column_dimensions['E'].width = 15.72

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = 9
    ws.page_setup.scale = 79

    ws.page_margins = PageMargins(
        left=2/2.54,
        right=2/2.54,
        top=4.5/2.54,
        bottom=2/2.54,
        header=0.8/2.54,
        footer=0.8/2.54
    )

    ws.print_options.horizontalCentered = True

    wb.save(nome_saida)

caminho_pdf = "RateioPeriodo_Report REF 03-2025.pdf"
caminho_excel_saidas = "Movimento do Caixa REF 03-2025.xlsx"

entradas = ler_entradas_pdf(caminho_pdf)
saidas = ler_saidas_excel(caminho_excel_saidas)

dados_unificados = juntar_movimentos_por_data(entradas, saidas)

gerar_livro_diario(dados_unificados, "LIVRO DIARIO - 03-25.xlsx", max_chars_saida=52)

print("criado, cuida")